#!/usr/bin/env python3
"""
run_baseline_test.py
Runs RBDtector on raw (unprocessed) EDF files to establish baseline results.

Tests: Test1.EDF, Test2.EDF
Purpose: Determine if preprocessing is necessary by checking if raw data produces valid results.

Author: Claude Code
Date: 2025-11-05
"""

import sys
import os
from pathlib import Path
import shutil
import pyedflib

# Add RBDtector to path
WORKSPACE = Path("/Users/hyeongsuk/Desktop/workspace/SNUH/Atonia_Index")
RBDTECTOR_PATH = WORKSPACE / "SW" / "RBDtector" / "RBDtector"
sys.path.insert(0, str(RBDTECTOR_PATH))

# Import RBDtector modules
from app_logic.PSG_controller import single_psg_run
from util import settings

# Import annotation converter
sys.path.insert(0, str(WORKSPACE / "SW" / "code"))
from convert_edf_annotations import convert_edf_annotations


def setup_test_directory(test_name, edf_source, work_dir):
    """
    Create test directory and copy EDF file.

    Args:
        test_name: Name of test (e.g., "Test1")
        edf_source: Source EDF file path
        work_dir: Working directory for processing

    Returns:
        Path to test directory
    """
    test_dir = work_dir / test_name
    test_dir.mkdir(parents=True, exist_ok=True)

    # Copy EDF file to test directory (use lowercase .edf for RBDtector compatibility)
    dest_edf = test_dir / f"{test_name}.edf"
    if not dest_edf.exists():
        print(f"Copying {edf_source.name} to {test_dir}/ (as {dest_edf.name})")
        shutil.copy2(edf_source, dest_edf)
    else:
        print(f"{dest_edf.name} already exists in {test_dir}/")

    return test_dir, dest_edf


def extract_annotations(edf_path):
    """
    Extract annotations from EDF file to text files.

    Args:
        edf_path: Path to EDF file

    Returns:
        True if successful, False otherwise
    """
    print(f"\nExtracting annotations from {edf_path.name}...")
    try:
        convert_edf_annotations(str(edf_path))
        return True
    except Exception as e:
        print(f"ERROR extracting annotations: {e}")
        return False


def configure_rbdtector_for_clinical_data():
    """
    Configure RBDtector settings for our clinical data.

    Clinical data has:
    - CHIN EMG: Channel 8 (EMG CHIN1-CHINz)
    - LEG EMG: Channel 12 (EMG RLEG+), Channel 13 (EMG LLEG+)
    - NO ARM EMG
    """
    print("\nConfiguring RBDtector for clinical data...")
    print("  Channels: CHIN + 2 LEGS (no arms)")

    # Configure channel names to match our EDF files
    settings.SIGNALS_TO_EVALUATE = ['EMG CHIN1-CHINz', 'EMG RLEG+', 'EMG LLEG+']
    settings.CHIN = 0  # Index 0 = EMG CHIN1-CHINz
    settings.LEGS = [1, 2]  # Indices 1, 2 = EMG RLEG+, EMG LLEG+
    settings.ARMS = []  # No arm channels

    # Keep other settings at default
    print(f"  SIGNALS_TO_EVALUATE: {settings.SIGNALS_TO_EVALUATE}")
    print(f"  CHIN: {settings.CHIN}")
    print(f"  LEGS: {settings.LEGS}")
    print(f"  ARMS: {settings.ARMS}")


def run_rbdtector_on_test(test_dir, test_name):
    """
    Run RBDtector analysis on a test directory.

    Args:
        test_dir: Path to test directory containing EDF and annotation files
        test_name: Name of test (for logging)

    Returns:
        dict with status and results
    """
    print(f"\n{'='*70}")
    print(f"Running RBDtector on {test_name}")
    print(f"{'='*70}")

    result = {
        'test_name': test_name,
        'test_dir': str(test_dir),
        'status': 'unknown',
        'error': None,
        'output_files': []
    }

    try:
        # Run RBDtector
        print(f"Processing: {test_dir}")
        single_psg_run(str(test_dir))

        # Check for output files
        output_dir = test_dir / "RBDtector output"
        if output_dir.exists():
            output_files = list(output_dir.glob("*"))
            result['output_files'] = [str(f) for f in output_files]

            if output_files:
                result['status'] = 'success'
                print(f"\n✅ SUCCESS: Generated {len(output_files)} output files")
                for f in output_files:
                    print(f"   - {f.name}")
            else:
                result['status'] = 'no_output'
                print(f"\n⚠️  WARNING: RBDtector output directory is empty")
        else:
            result['status'] = 'no_output_dir'
            print(f"\n⚠️  WARNING: No RBDtector output directory created")

    except Exception as e:
        result['status'] = 'error'
        result['error'] = str(e)
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()

    return result


def inspect_results(result):
    """
    Inspect RBDtector results and check for common issues.

    Args:
        result: Result dictionary from run_rbdtector_on_test
    """
    print(f"\n{'-'*70}")
    print(f"Result Inspection: {result['test_name']}")
    print(f"{'-'*70}")

    if result['status'] == 'success':
        # Look for Excel file and check if it contains data
        excel_files = [f for f in result['output_files'] if f.endswith('.xlsx')]

        if excel_files:
            import openpyxl
            for excel_file in excel_files:
                print(f"\nInspecting: {Path(excel_file).name}")
                try:
                    wb = openpyxl.load_workbook(excel_file, data_only=True)
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        rows = list(sheet.iter_rows(values_only=True))

                        print(f"  Sheet: {sheet_name}")
                        print(f"    Rows: {len(rows)}")

                        # Check if there's actual data (not just headers)
                        if len(rows) > 1:
                            # Look for numeric values in the data
                            data_rows = rows[1:]  # Skip header
                            has_nonzero = False
                            has_null = False

                            for row in data_rows:
                                for cell in row:
                                    if isinstance(cell, (int, float)):
                                        if cell != 0:
                                            has_nonzero = True
                                        if cell is None:
                                            has_null = True

                            if not has_nonzero:
                                print(f"    ⚠️  WARNING: All numeric values are 0 or null!")
                            else:
                                print(f"    ✅ Contains non-zero data")
                        else:
                            print(f"    ⚠️  WARNING: Only header row, no data!")

                    wb.close()
                except Exception as e:
                    print(f"  ERROR reading Excel file: {e}")
    else:
        print(f"Status: {result['status']}")
        if result['error']:
            print(f"Error: {result['error']}")


def main():
    """Main execution function."""
    print("="*70)
    print("RBDtector Baseline Test / RBDtector 기준선 테스트")
    print("="*70)
    print("Purpose: Test RBDtector on raw (unprocessed) data")
    print("목적: 전처리하지 않은 원본 데이터로 RBDtector 테스트")
    print("="*70)

    # Paths
    clinical_db = WORKSPACE / "Clinical_DB"
    work_dir = WORKSPACE / "Results" / "raw"
    work_dir.mkdir(parents=True, exist_ok=True)

    # Test files
    test_files = [
        ("Test1", clinical_db / "Test1.EDF"),
        ("Test2", clinical_db / "Test2.EDF")
    ]

    # Configure RBDtector
    configure_rbdtector_for_clinical_data()

    # Process each test file
    results = []
    for test_name, edf_source in test_files:
        print(f"\n{'#'*70}")
        print(f"# Processing: {test_name}")
        print(f"{'#'*70}")

        # Setup directory
        test_dir, dest_edf = setup_test_directory(test_name, edf_source, work_dir)

        # Extract annotations
        if not extract_annotations(dest_edf):
            print(f"❌ Failed to extract annotations for {test_name}")
            continue

        # Run RBDtector
        result = run_rbdtector_on_test(test_dir, test_name)
        results.append(result)

        # Inspect results
        inspect_results(result)

    # Summary
    print(f"\n{'='*70}")
    print("BASELINE TEST SUMMARY / 기준선 테스트 요약")
    print(f"{'='*70}")

    for result in results:
        status_symbol = {
            'success': '✅',
            'no_output': '⚠️ ',
            'no_output_dir': '⚠️ ',
            'error': '❌'
        }.get(result['status'], '❓')

        print(f"{status_symbol} {result['test_name']}: {result['status']}")
        if result['output_files']:
            print(f"   Output files: {len(result['output_files'])}")

    print(f"\n{'='*70}")
    print("Next step: Document these baseline results")
    print("다음 단계: 이 기준선 결과를 문서화")
    print(f"{'='*70}")

    return results


if __name__ == "__main__":
    results = main()
